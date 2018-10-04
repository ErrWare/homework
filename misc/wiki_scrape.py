import json
import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import logging
import os
import re
logging.basicConfig(level=logging.CRITICAL)


quant_re = re.compile('([0-9]+\.?[0-9]*)(\s.?g)?')
percent_re = re.compile('[0-9]+%')

with open('wiki_scrape_info.json','r') as info_file:
	info = json.load(info_file)

WORKBOOK_NAME = 'wiki_scrape2.xlsx'
NUTRITION_SHEET_NAME = 'Nutrition Table'

if WORKBOOK_NAME in os.listdir():
	workbook = xl.load_workbook(WORKBOOK_NAME)
else:
	workbook = xl.Workbook()

logging.info('active sheet: ' + workbook.active.title)
if 'Sheet' in workbook:
	scrape_sheet = workbook['Sheet']
	scrape_sheet.title = 'Nutrition Table'
else:
	scrape_sheet = workbook['Nutrition Table']

row_num = 1
# First Col - caption
# Second Col - empty for now
# Third Col - ... info
scrape_sheet['A1'].value = 'Name'
offset = 3
for position, col_name in enumerate(info['info']):
	scrape_sheet.cell(row=1,column=position+offset).value = col_name
row_num += 1


for food in info['food_articles']:
	res = requests.get(food)
	soup = BeautifulSoup(res.text, features='html5lib')
	tables = soup.find_all('table', class_='infobox nowrap')

	print(food, '\ttables found: ', len(tables))
	for table in tables:
		scrape_sheet.cell(row=row_num,column=1).value = table.caption.text
		# it's unpythonic I know, but variable tables / article
		for position, col_name in enumerate(info['info']):
			try:
				a_tag = table.find('a',text=lambda t: t is not None and t.lower() == col_name)
				if a_tag is None:
					logging.info('None on <{}, {}>'.format(food[-12:],col_name))
					continue
				tr_tag = a_tag.find_parent('tr')
				td_tag = tr_tag.find('td')
				text = percent_re.sub('',td_tag.text).strip()
				#div_tag = td_tag.find('div')
				#logging.info(div_tag.text)
				#quantity = re.search('.*\s', div_tag.text)
				
				quant = quant_re.search(text)
				if quant is None:
					logging.info('{}:: QUANT IS NONE on text "{}"'.format(col_name,text))
				else:
					logging.info('{}:: quant found: {}'.format(col_name,quant.group(0)))
				logging.info('Writing {} to cell {}'.format(quant.group(1), col_name))
				scrape_sheet.cell(row=row_num,column=position+offset).value = '='+quant.group(1)#'='+quantity
			except Exception as e:
				logging.info(col_name + '\t' + str(e))
				
		row_num += 1





#if 'DietPlanner' not in workbook:
logging.info('making planner')
planner_sheet = workbook.create_sheet('DietPlanner')
planner_sheet['A1'].value = 'Name'
planner_sheet['B1'].value = 'Amount'

for position, col_name in enumerate(info['info']):
	planner_sheet.cell(row=1,column=position+offset).value = col_name
for row in range(2,row_num):
	planner_sheet.cell(row=row,column=1).value = workbook[NUTRITION_SHEET_NAME].cell(row=row,column=1).value
	for position, _ in enumerate(info['info']):
		formula = (
			"=B{} * '{}'!{}{} / 100".format(row,NUTRITION_SHEET_NAME,
			xl.utils.get_column_letter(position+offset),row)
			)
		
		logging.info(formula)
		planner_sheet.cell(row=row,column=position+offset).value = formula
	
for position, _ in enumerate(info['info']):
	col_char = xl.utils.get_column_letter(position+offset)
	formula = '=SUM({0}2:{0}{1})'.format(col_char,row_num-1)
	logging.info(formula)
	planner_sheet.cell(row=row_num, column=position+offset).value = formula

workbook.save(WORKBOOK_NAME)
workbook.close()